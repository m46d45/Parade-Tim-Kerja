"""
Parade of Trades – Analysis: replications, statistics, export
=============================================================

Supports:
  - Multiple independent replications with summary statistics
  - Side-by-side scenario comparison (incl. Tommelein 2020 trio)
  - Export single-run and multi-replication results to CSV / Excel

Usage
-----
>>> from parade_of_trades_core import ParadeConfig
>>> from parade_of_trades_analysis import run_replications, export_result_excel
>>> cfg = ParadeConfig.from_preset("medium")
>>> batch = run_replications(cfg, n_reps=100, seed_base=0)
>>> print(batch.summary_table())
>>> export_result_excel(batch.results[0], "run.xlsx")
>>> batch.export_excel("reps.xlsx")
"""

from __future__ import annotations

import csv
import re
import math
import statistics
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple, Union

from parade_of_trades_core import (
    ParadeConfig,
    ParadeOfTrades,
    ParadeResult,
    tommelein2020_scenarios,
)


PathLike = Union[str, Path]


# ---------------------------------------------------------------------------
# Little's Law (classic production form; yield y = 1 in this app)
# ---------------------------------------------------------------------------
# Reference: WIP = CT × TH  (Little; Hopp & Spearman / Factory Physics;
# Project Production Institute — "Little's Law in Production Systems with Yield Loss")
# Our parade model has no scrap/yield loss (y_i = 1), so classic form applies.


@dataclass
class LittlesLawMetrics:
    """Time-average Little's Law metrics from a completed parade run."""

    throughput: float
    """System throughput TH (zona selesai proyek / periode) = N / duration."""

    avg_pipeline_wip: float
    """Average pipeline WIP: zona yang sudah dikerjakan T1 tetapi belum selesai di T5."""

    avg_buffer_wip: float
    """Average interface buffer WIP (jumlah semua buffer antar-tim)."""

    peak_pipeline_wip: float
    peak_buffer_wip: float

    cycle_time_pipeline: float
    """CT from Little: avg_pipeline_wip / TH (periode)."""

    cycle_time_buffer: float
    """CT if only buffer WIP is used: avg_buffer_wip / TH."""

    duration: int
    total_units: int
    yield_assumed: float = 1.0
    """App model has no yield loss; Y = 1 so TH_end = TH_start."""

    check_pipeline: float = 0.0
    """TH × CT_pipeline should ≈ avg_pipeline_wip."""

    check_buffer: float = 0.0

    def as_rows(self) -> List[dict]:
        return [
            {
                "Metrik": "Throughput (TH)",
                "Nilai": round(self.throughput, 4),
                "Satuan": "zona / periode",
                "Keterangan": "N ÷ durasi proyek",
            },
            {
                "Metrik": "WIP pipeline (rata-rata)",
                "Nilai": round(self.avg_pipeline_wip, 3),
                "Satuan": "zona",
                "Keterangan": "Rata-rata (kumulatif T1 − kumulatif T5)",
            },
            {
                "Metrik": "WIP buffer (rata-rata)",
                "Nilai": round(self.avg_buffer_wip, 3),
                "Satuan": "zona",
                "Keterangan": "Rata-rata jumlah semua buffer antar-tim",
            },
            {
                "Metrik": "CT pipeline (Little)",
                "Nilai": round(self.cycle_time_pipeline, 3),
                "Satuan": "periode",
                "Keterangan": "WIP_pipeline ÷ TH",
            },
            {
                "Metrik": "CT buffer (Little)",
                "Nilai": round(self.cycle_time_buffer, 3),
                "Satuan": "periode",
                "Keterangan": "WIP_buffer ÷ TH (hanya antrian antar-tim)",
            },
            {
                "Metrik": "WIP puncak pipeline",
                "Nilai": round(self.peak_pipeline_wip, 3),
                "Satuan": "zona",
                "Keterangan": "Maks (T1 − T5) kumulatif",
            },
            {
                "Metrik": "WIP puncak buffer",
                "Nilai": round(self.peak_buffer_wip, 3),
                "Satuan": "zona",
                "Keterangan": "Maks total buffer serentak",
            },
            {
                "Metrik": "Cek Little (pipeline)",
                "Nilai": round(self.check_pipeline, 3),
                "Satuan": "zona",
                "Keterangan": "TH × CT ≈ WIP rata-rata (harus dekat)",
            },
        ]


def littles_law_metrics(result: ParadeResult) -> LittlesLawMetrics:
    """
    Compute Little's Law metrics from a parade result.

    Classic form (no yield loss):  **WIP = TH × CT**

    - **TH** = total_units / duration (average exit rate of finished zones).
    - **WIP pipeline** at period t ≈ cumulative_T1(t) − cumulative_T5(t)
      (zona yang sudah masuk jalur tetapi belum keluar di finishing).
    - **WIP buffer** = sum of interface buffers (antrian antar-tim saja).
    - **CT** = WIP / TH  (waktu tinggal rata-rata implisit).

    Yield loss (artikel PPI): model app ini y_i = 1 untuk semua tahap, jadi
    TH_end = TH_start dan bentuk klasik berlaku tanpa koreksi yield.
    """
    n = result.config.n_trades
    total = result.config.total_units
    duration = max(1, int(result.duration))
    th = float(total) / float(duration)

    pipeline_series: List[float] = []
    buffer_series: List[float] = []

    # period 0 empty
    pipeline_series.append(0.0)
    buffer_series.append(0.0)

    for rec in result.history:
        if rec.cumulative and len(rec.cumulative) >= n:
            pipe = float(rec.cumulative[0] - rec.cumulative[-1])
            pipeline_series.append(max(0.0, pipe))
        else:
            pipeline_series.append(0.0)
        if rec.buffers:
            buffer_series.append(float(sum(rec.buffers)))
        else:
            buffer_series.append(0.0)

    avg_pipe = sum(pipeline_series) / len(pipeline_series)
    avg_buf = sum(buffer_series) / len(buffer_series)
    peak_pipe = max(pipeline_series) if pipeline_series else 0.0
    peak_buf = max(buffer_series) if buffer_series else 0.0

    ct_pipe = avg_pipe / th if th > 0 else float("inf")
    ct_buf = avg_buf / th if th > 0 else float("inf")

    return LittlesLawMetrics(
        throughput=th,
        avg_pipeline_wip=avg_pipe,
        avg_buffer_wip=avg_buf,
        peak_pipeline_wip=peak_pipe,
        peak_buffer_wip=peak_buf,
        cycle_time_pipeline=ct_pipe,
        cycle_time_buffer=ct_buf,
        duration=duration,
        total_units=total,
        yield_assumed=1.0,
        check_pipeline=th * ct_pipe if th > 0 and math.isfinite(ct_pipe) else 0.0,
        check_buffer=th * ct_buf if th > 0 and math.isfinite(ct_buf) else 0.0,
    )


def littles_law_series(result: ParadeResult) -> Dict[str, List[float]]:
    """Time series for plotting pipeline WIP and buffer WIP."""
    n = result.config.n_trades
    pipe = [0.0]
    buf = [0.0]
    for rec in result.history:
        if rec.cumulative and len(rec.cumulative) >= n:
            pipe.append(max(0.0, float(rec.cumulative[0] - rec.cumulative[-1])))
        else:
            pipe.append(0.0)
        buf.append(float(sum(rec.buffers)) if rec.buffers else 0.0)
    return {"pipeline_wip": pipe, "buffer_wip": buf, "period": list(range(len(pipe)))}




# ---------------------------------------------------------------------------
# Kingman's Equation (VUT approximation, Factory Physics)
# ---------------------------------------------------------------------------
# CT ≈ t_e + ((c_a² + c_e²)/2) × (u/(1-u)) × t_e
#    = V × U × T  form: wait ≈ V×U×t_e, CT = wait + t_e
# c_a = CV arrivals, c_e = CV process (effective) time, u = utilization, t_e = mean process time
# Ref: Hopp & Spearman, Factory Physics; Kingman (1961) G/G/1 heavy-traffic approx.


@dataclass
class KingmanStationRow:
    trade_index: int
    name: str
    utilization: float
    t_e: float
    """Mean process time per zona (periode)."""
    c_e: float
    """CV of process time (from capacity variability config)."""
    c_a: float
    """CV of arrivals (0 for T1; else ≈ c_e upstream)."""
    wait_kingman: float
    ct_kingman: float
    ct_observed: float
    """Empirical: time_on_site / production (periode per zona di stasiun)."""
    v_factor: float
    u_factor: float


@dataclass
class KingmanMetrics:
    stations: List[KingmanStationRow]
    sum_ct_kingman: float
    sum_ct_observed: float
    system_ct_little: float
    """Little's Law pipeline CT for comparison."""
    bottleneck_u: float
    note: str = ""

    def as_rows(self) -> List[dict]:
        rows = []
        for s in self.stations:
            rows.append({
                "Tim": f"T{s.trade_index + 1} {s.name}",
                "u": round(s.utilization, 3),
                "t_e": round(s.t_e, 3),
                "c_e": round(s.c_e, 3),
                "c_a": round(s.c_a, 3),
                "V=(c_a²+c_e²)/2": round(s.v_factor, 3),
                "U=u/(1-u)": round(s.u_factor, 3) if math.isfinite(s.u_factor) else "∞",
                "Wait Kingman": round(s.wait_kingman, 3) if math.isfinite(s.wait_kingman) else "∞",
                "CT Kingman": round(s.ct_kingman, 3) if math.isfinite(s.ct_kingman) else "∞",
                "CT amati": round(s.ct_observed, 3),
            })
        return rows


def _process_time_moments(trade) -> Tuple[float, float]:
    """
    Mean and CV of process time per zone from capacity model.

    Capacity C = zona/periode → process time T = 1/C periode/zona.
    """
    lo = max(float(trade.low), 1e-12)
    hi = max(float(trade.high), 1e-12)
    p = float(trade.p_high)
    if getattr(trade, "deterministic", False) or abs(lo - hi) < 1e-12:
        base = float(trade.base_speed) if trade.base_speed is not None else float(trade.mean)
        base = max(base, 1e-12)
        return 1.0 / base, 0.0
    t_lo, t_hi = 1.0 / lo, 1.0 / hi
    t_e = (1.0 - p) * t_lo + p * t_hi
    var = (1.0 - p) * (t_lo - t_e) ** 2 + p * (t_hi - t_e) ** 2
    c_e = math.sqrt(max(var, 0.0)) / t_e if t_e > 0 else 0.0
    return t_e, c_e


def kingman_ct(u: float, t_e: float, c_a: float, c_e: float) -> Tuple[float, float, float, float]:
    """
    Returns (wait, ct, v_factor, u_factor).

    Kingman / VUT: wait ≈ ((c_a² + c_e²)/2) * (u/(1-u)) * t_e
                   CT  ≈ wait + t_e
    """
    u = float(u)
    t_e = max(float(t_e), 1e-12)
    c_a = max(float(c_a), 0.0)
    c_e = max(float(c_e), 0.0)
    v = 0.5 * (c_a ** 2 + c_e ** 2)
    if u >= 1.0 - 1e-9:
        return float("inf"), float("inf"), v, float("inf")
    if u <= 0.0:
        return 0.0, t_e, v, 0.0
    u_factor = u / (1.0 - u)
    wait = v * u_factor * t_e
    return wait, wait + t_e, v, u_factor


def kingman_metrics(result: ParadeResult) -> KingmanMetrics:
    """
    Per-trade Kingman (VUT) approximation vs observed station cycle time.

    - **t_e, c_e** from trade capacity / variability configuration.
    - **u** from simulated utilization (production / effective capacity).
    - **c_a**: T1 ≈ 0 (pasokan zona mentah dianggap teratur); hilir ≈ **c_e**
      stasiun hulu (pendekatan tandem teaching model).
    - **CT amati** ≈ time_on_site / production (periode per zona di stasiun).

    Bandingkan Σ CT Kingman dengan CT pipeline Little's Law (orde yang sama,
    tidak harus sama — Kingman stasioner vs proyek berhingga).
    """
    stations: List[KingmanStationRow] = []
    prev_c_e = 0.0
    for i, m in enumerate(result.trade_metrics):
        trade = result.config.trades[i]
        t_e, c_e = _process_time_moments(trade)
        u = float(m.utilization)
        # Cap u slightly below 1 for display formula stability when util≈1
        u_calc = min(u, 0.999)
        c_a = 0.0 if i == 0 else prev_c_e
        wait, ct_k, v_f, u_f = kingman_ct(u_calc, t_e, c_a, c_e)
        prod = max(int(m.total_production), 1)
        ct_obs = float(m.time_on_site) / float(prod)
        stations.append(
            KingmanStationRow(
                trade_index=i,
                name=m.name,
                utilization=u,
                t_e=t_e,
                c_e=c_e,
                c_a=c_a,
                wait_kingman=wait,
                ct_kingman=ct_k,
                ct_observed=ct_obs,
                v_factor=v_f,
                u_factor=u_f,
            )
        )
        prev_c_e = c_e

    ll = littles_law_metrics(result)
    sum_k = sum(s.ct_kingman for s in stations if math.isfinite(s.ct_kingman))
    sum_o = sum(s.ct_observed for s in stations)
    bott_u = max((s.utilization for s in stations), default=0.0)
    note = (
        "Kingman = pendekatan antrian stasioner (VUT). "
        "Proyek parade berhingga + batch handoff bisa beda dari prediksi; "
        "pakai untuk intuisi: V↑ atau U↑ → CT↑."
    )
    return KingmanMetrics(
        stations=stations,
        sum_ct_kingman=sum_k,
        sum_ct_observed=sum_o,
        system_ct_little=ll.cycle_time_pipeline,
        bottleneck_u=bott_u,
        note=note,
    )


def kingman_combined(result: ParadeResult) -> dict:
    """Aggregate (gabungan) Kingman inputs/outputs for CT–u chart annotation."""
    k = kingman_metrics(result)
    n = max(len(k.stations), 1)
    u_bar = sum(s.utilization for s in k.stations) / n
    t_e = sum(s.t_e for s in k.stations) / n
    c_a = sum(s.c_a for s in k.stations) / n
    c_e = sum(s.c_e for s in k.stations) / n
    v = 0.5 * (c_a ** 2 + c_e ** 2)
    u_c = min(u_bar, 0.999)
    wait, ct, _, u_f = kingman_ct(u_c, t_e, c_a, c_e)
    return {
        "u_bar": u_bar,
        "t_e": t_e,
        "c_a": c_a,
        "c_e": c_e,
        "v": v,
        "u_factor": u_f,
        "wait": wait,
        "ct": ct,
        "bottleneck_u": k.bottleneck_u,
    }




def _interp_xy(xs: Sequence[float], ys: Sequence[float], x: float) -> float:
    """Linear interpolation; clamp to ends."""
    if not xs:
        return float("nan")
    if x <= xs[0]:
        return float(ys[0])
    if x >= xs[-1]:
        return float(ys[-1])
    for i in range(1, len(xs)):
        if x <= xs[i]:
            x0, x1 = xs[i - 1], xs[i]
            y0, y1 = ys[i - 1], ys[i]
            if abs(x1 - x0) < 1e-15:
                return float(y0)
            t = (x - x0) / (x1 - x0)
            return float(y0 + t * (y1 - y0))
    return float(ys[-1])


def evaluate_at_wip(result: ParadeResult, wip_level: float) -> Dict[str, float]:
    """
    Predicted TH & CT if system operated at given WIP (e.g. CONWIP limit).

    Uses the actual (var) operations curve; also reports best-case bounds
    and delta vs current operating point.
    """
    d = littles_operations_curve(result)
    w = max(float(wip_level), 1e-9)

    # sort actual
    aw = list(d["wip"])
    ath = list(d["th"])
    act = list(d["ct"])
    if aw:
        order = sorted(range(len(aw)), key=lambda i: aw[i])
        aw = [aw[i] for i in order]
        ath = [ath[i] for i in order]
        act = [act[i] for i in order]
    th_a = _interp_xy(aw, ath, w) if aw else float("nan")
    ct_a = _interp_xy(aw, act, w) if aw else float("nan")
    # consistency via Little if one missing
    if math.isfinite(th_a) and th_a > 1e-12 and (not math.isfinite(ct_a) or ct_a <= 0):
        ct_a = w / th_a
    if math.isfinite(ct_a) and ct_a > 1e-12 and (not math.isfinite(th_a) or th_a <= 0):
        th_a = w / ct_a

    bw = list(d.get("bc_wip") or [])
    bth = list(d.get("bc_th") or [])
    bct = list(d.get("bc_ct") or [])
    th_b = _interp_xy(bw, bth, w) if bw else float("nan")
    ct_b = _interp_xy(bw, bct, w) if bw else float("nan")

    op_w = float(d["op_wip"])
    op_th = float(d["op_th"])
    op_ct = float(d["op_ct"])

    return {
        "wip": w,
        "th": th_a,
        "ct": ct_a,
        "th_best": th_b,
        "ct_best": ct_b,
        "op_wip": op_w,
        "op_th": op_th,
        "op_ct": op_ct,
        "d_th": th_a - op_th,
        "d_ct": ct_a - op_ct,
        "d_wip": w - op_w,
        "w_min": float(d["w_min"]),
        "w_opt": float(d["w_opt"]),
        "th_max": float(d["th_max"]),
    }


def littles_operations_curve(
    result: ParadeResult,
    n_points: int = 80,
) -> Dict[str, List[float]]:
    """
    Operations curves linking WIP–TH–CT (Little + Kingman + best-case bound).

    **Best case (Factory Physics, no variability, bottleneck capacity):**
      TH_max = min_i mean capacity (zona/periode)
      T0     = sum_i t_e,i   (raw process time through all stations)
      W0     = TH_max × T0   (critical WIP)
      For W ≤ W0:  TH = W/T0 ,  CT = T0
      For W ≥ W0:  TH = TH_max , CT = W/TH_max

    **Actual (with variability):** Kingman CT(u), TH≈u/t_e_avg, WIP=TH×CT
    — lies *worse* than the best-case envelope (lower TH / higher CT).
    """
    comb = kingman_combined(result)
    t_e_avg = max(float(comb["t_e"]), 1e-9)
    c_a = float(comb["c_a"])
    c_e = float(comb["c_e"])

    # --- Best-case parameters from config (no var, max capacity = means) ---
    means = []
    t0 = 0.0
    for tr in result.config.trades:
        mean_c = float(tr.mean) if tr.mean and tr.mean > 0 else max(float(tr.base_speed or 1.0), 1e-9)
        means.append(mean_c)
        t0 += 1.0 / max(mean_c, 1e-9)
    th_max = min(means) if means else 1.0
    t0 = max(t0, 1e-9)
    w0 = th_max * t0  # critical WIP

    ll = littles_law_metrics(result)
    V_early = max(float(comb.get("v") or 0.0), 0.0)
    # Placeholder W_opt estimate for axis range (refined below)
    alpha_tmp = 0.90
    if V_early < 1e-9:
        w_opt_est = w0
    else:
        w_opt_est = alpha_tmp * w0 * (1.0 + V_early * (alpha_tmp / (1.0 - alpha_tmp)))
        w_opt_est = max(w_opt_est, w0)

    # Shared WIP axis: only a little past W_opt / operating WIP (+5 WIP)
    w_hi = max(w0, w_opt_est, float(ll.avg_pipeline_wip), 1.0) + 5.0
    w_grid = [max(w_hi * i / max(n_points - 1, 1), 1e-6) for i in range(n_points)]

    # Best-case envelope (extended full range)
    bc_th: List[float] = []
    bc_ct: List[float] = []
    for w in w_grid:
        if w <= w0 + 1e-12:
            bc_th.append(w / t0)
            bc_ct.append(t0)
        else:
            bc_th.append(th_max)
            bc_ct.append(w / th_max)

    # Actual curve: Kingman parametric in u, then extend by Little's Law
    # at saturated TH so curves continue past W_opt / operating WIP.
    wips: List[float] = []
    ths: List[float] = []
    cts: List[float] = []
    us: List[float] = []
    n_u = max(n_points, 120)
    for i in range(n_u):
        # push utilization close to 1 so CT/WIP grow
        u = 0.02 + (0.985 - 0.02) * i / max(n_u - 1, 1)
        wait, ct, _, _ = kingman_ct(u, t_e_avg, c_a, c_e)
        if not math.isfinite(ct) or ct <= 0:
            continue
        th = min(u / t_e_avg, th_max * 0.999)
        if th <= 1e-12:
            continue
        wip = th * ct
        us.append(u)
        ths.append(th)
        cts.append(ct)
        wips.append(wip)

    # Sort by WIP and extend to w_hi if needed (CT = WIP/TH, TH → th_sat)
    if wips:
        order = sorted(range(len(wips)), key=lambda i: wips[i])
        wips = [wips[i] for i in order]
        ths = [ths[i] for i in order]
        cts = [cts[i] for i in order]
        us = [us[i] for i in order]
        w_last = wips[-1]
        th_sat = ths[-1]
        # gently approach a practical ceiling ≤ th_max
        th_ceil = min(th_max * (0.92 if V_early > 1e-9 else 0.999), max(th_sat, th_max * 0.5))
        if w_last < w_hi - 1e-6:
            n_ext = max(20, n_points // 2)
            for j in range(1, n_ext + 1):
                w = w_last + (w_hi - w_last) * j / n_ext
                # TH rises slightly toward ceiling then flattens
                frac = j / n_ext
                th = th_sat + (th_ceil - th_sat) * (1.0 - math.exp(-3.0 * frac))
                th = min(th, th_max * 0.999)
                ct = w / th  # Little's Law extension
                wips.append(w)
                ths.append(th)
                cts.append(ct)
                us.append(us[-1] if us else 0.9)

    # --- WIP landmarks (project production / Factory Physics) ---
    # W_min = W0 = critical WIP (best-case minimum WIP to reach TH_max).
    #   W0 = TH_max × T0
    w_min = float(w0)

    # W_opt = practical optimal WIP under variability (Hopp & Spearman style):
    #   Without variability (V≈0): W_opt = W_min  (deterministic critical is enough)
    #   With variability: need more WIP to hold high utilization / TH.
    #   Approximate at target utilization α:
    #     inflation = 1 + V × α/(1-α)     (VUT wait factor on raw process)
    #     W_opt = α × W0 × inflation
    #   so W_opt > W_min whenever V > 0.
    V = max(float(comb.get("v") or 0.0), 0.0)
    alpha = 0.90  # target fraction of bottleneck rate for "optimal" region
    if V < 1e-9:
        w_opt = w_min
    else:
        inflation = 1.0 + V * (alpha / (1.0 - alpha))
        w_opt = alpha * w_min * inflation
        # never below critical
        w_opt = max(w_opt, w_min)

    # Practical upper teaching mark: WIP if pushing α→0.95 with same V
    alpha_hi = 0.95
    if V < 1e-9:
        w_pwc = w_min  # same as critical when deterministic
    else:
        infl_hi = 1.0 + V * (alpha_hi / (1.0 - alpha_hi))
        w_pwc = max(w_opt, alpha_hi * w_min * infl_hi)

    # CONWIP suggestion = W_opt (cap WIP near optimal)
    conwip = float(w_opt)

    return {
        "wip": wips,
        "th": ths,
        "ct": cts,
        "u": us,
        "op_wip": ll.avg_pipeline_wip,
        "op_th": ll.throughput,
        "op_ct": ll.cycle_time_pipeline,
        "t_e": t_e_avg,
        "v": float(comb["v"]),
        "u_bar": float(comb["u_bar"]),
        # best-case envelope
        "bc_wip": w_grid,
        "bc_th": bc_th,
        "bc_ct": bc_ct,
        "th_max": th_max,
        "t0": t0,
        "w0": w0,
        "w_min": w_min,
        "w_opt": float(w_opt),
        "w_pwc": float(w_pwc),
        "v_factor": float(V),
        "conwip": conwip,
    }



def _unit_normal_loss(z: float) -> float:
    """Unit normal loss G(z) = φ(z) - z (1-Φ(z))."""
    # φ and Φ without scipy
    phi = math.exp(-0.5 * z * z) / math.sqrt(2.0 * math.pi)
    # Abramowitz-Stegun approximation for Φ
    t = 1.0 / (1.0 + 0.2316419 * abs(z))
    d = 0.3989423 * math.exp(-0.5 * z * z)
    p = d * t * (0.3193815 + t * (-0.3565638 + t * (1.781478 + t * (-1.821256 + t * 1.330274))))
    Phi = 1.0 - p if z >= 0 else p
    return phi - z * (1.0 - Phi)


def inventory_fill_rate_metrics(result: ParadeResult) -> dict:
    """
    Inventory vs fill-rate view of the parade.

    **Inventory** = average interface buffer WIP (zona menunggu di antara tim).
    **Fill rate** (tipe-2, analog) per tim hilir ≈
        produksi / (produksi + idle)
    yaitu fraksi kapasitas yang benar-benar menjadi output (bukan starvation).

    Juga kurva teoritis base-stock (normal loss) untuk konteks Factory Physics:
        safety stock ~ z·σ, fill rate naik cekung menuju 100% saat inventory naik.
    """
    n = result.config.n_trades
    # Per-interface inventory
    buf = result.buffer_series()  # [iface][period]
    iface_rows = []
    for j in range(result.config.n_interfaces):
        series = buf[j] if j < len(buf) else []
        avg_inv = sum(series) / len(series) if series else 0.0
        peak = max(series) if series else 0.0
        up = result.config.trades[j].name
        down = result.config.trades[j + 1].name
        # fill rate of downstream trade
        m = result.trade_metrics[j + 1]
        denom = m.total_production + m.total_idle
        fr = (m.total_production / denom) if denom > 0 else 1.0
        iface_rows.append({
            "buffer": f"B{j + 1}",
            "from": up,
            "to": down,
            "avg_inventory": avg_inv,
            "peak_inventory": peak,
            "fill_rate": fr,
            "downstream_idle": m.total_idle,
            "downstream_prod": m.total_production,
        })

    # System aggregates
    avg_inv_sys = sum(r["avg_inventory"] for r in iface_rows) / max(len(iface_rows), 1)
    # Combined fill rate: production-weighted across T2..T5
    prod = sum(result.trade_metrics[i].total_production for i in range(1, n))
    idle = sum(result.trade_metrics[i].total_idle for i in range(1, n))
    fr_sys = prod / (prod + idle) if (prod + idle) > 0 else 1.0
    # T1 often ~1 if raw infinite
    m1 = result.trade_metrics[0]
    d1 = m1.total_production + m1.total_idle
    fr_t1 = m1.total_production / d1 if d1 > 0 else 1.0

    return {
        "interfaces": iface_rows,
        "avg_inventory_system": avg_inv_sys,
        "fill_rate_system": fr_sys,
        "fill_rate_t1": fr_t1,
        "peak_buffer_total": max(
            (sum(result.history[t].buffers) for t in range(len(result.history))),
            default=0,
        ) if result.history else 0,
    }


def inventory_fill_rate_curve(
    result: ParadeResult,
    n_points: int = 60,
) -> Dict[str, List[float]]:
    """
    Theoretical inventory–fill-rate curve (base-stock / normal demand).

    Parameterize by safety factor z:
      inventory_index = max(z, 0) * σ_proxy + cycle_stock
      fill_rate ≈ 1 - G(z) / (μ/σ)   clipped to [0,1]
    Shape: concave, diminishing returns of inventory on service.
    σ_proxy scaled from run variability (c_e).
    """
    comb = kingman_combined(result)
    # μ demand rate ~ TH; σ from CV ~ c_e * scale
    ll = littles_law_metrics(result)
    mu = max(ll.throughput, 0.1)
    ce = max(float(comb["c_e"]), 0.05)  # minimum for visible curve
    sigma = max(ce * mu, 0.05)

    invs: List[float] = []
    frs: List[float] = []
    # z from -0.5 to 3.2
    for i in range(n_points):
        z = -0.5 + (3.2 - (-0.5)) * i / max(n_points - 1, 1)
        # cycle stock proxy + safety
        cycle = mu * float(comb["t_e"]) * 0.5  # rough half process
        safety = max(z, 0.0) * sigma
        inv = cycle + safety
        # Type-2 style fill rate using unit normal loss
        # FR ≈ 1 - σ*G(z) / μ   (per period demand μ)
        G = _unit_normal_loss(z)
        fr = 1.0 - (sigma * G) / mu
        fr = max(0.0, min(1.0, fr))
        invs.append(max(inv, 0.0))
        frs.append(fr)

    emp = inventory_fill_rate_metrics(result)
    return {
        "inventory": invs,
        "fill_rate": frs,
        "op_inventory": emp["avg_inventory_system"],
        "op_fill_rate": emp["fill_rate_system"],
        "interfaces": emp["interfaces"],
        "mu": mu,
        "sigma": sigma,
    }



@dataclass
class TaktPlanCell:
    trade_index: int
    zone: int  # 1-based zone index
    period_start: int
    period_end: int  # inclusive
    planned_rate: float


@dataclass
class TaktPlan:
    """Ideal takt plan for zone-flow floor cycle (no variability)."""

    n_trades: int
    n_zones: int
    batch_size: int
    rate: float  # zones per period per trade (capacity)
    cells: List[TaktPlanCell]
    duration: float
    takt_time: float
    """Periods to complete one zone at one station (= 1/rate)."""
    handoff_lag: int = 1  # next-period release

    def as_rows(self) -> List[dict]:
        rows = []
        for c in self.cells:
            rows.append({
                "Tim": c.trade_index + 1,
                "Zona": c.zone,
                "Mulai": c.period_start,
                "Selesai": c.period_end,
                "Durasi": round(float(c.period_end) - float(c.period_start), 4),
            })
        return rows


def zone_rate_for_work(base_rate: float, n_zones: int, total_work: float) -> float:
    """Kapasitas zona/periode agar total kerja tetap: base_rate * n_zones / total_work."""
    return max(float(base_rate), 1e-9) * max(1, int(n_zones)) / max(float(total_work), 1e-9)


def build_takt_plan(
    n_trades: int = 5,
    n_zones: int = 20,
    batch_size: int = 4,
    rate: float = 1.0,
    handoff_lag: int = 1,
    total_work: Optional[float] = None,
) -> TaktPlan:
    """
    Ideal OPF takt train (period 0-based, continuous time).

    **Fixed total work (educational zoning):**
    - ``total_work`` = lingkup proyek tetap (default = n_zones → unit size 1)
    - Kapasitas ``rate`` = unit kerja / periode
    - Zona lebih banyak → ukuran zona = total_work/n_zones lebih kecil
      → waktu per zona lebih pendek → **durasi total lebih cepat**
      (pipeline fill (n_trades-1)×t_zona mengecil)

    Example: total_work=40, rate=1, 5 trades:
      30 zona → t_zona=40/30, T≈45.3 p
      40 zona → t_zona=1,     T≈44 p
      50 zona → t_zona=0.8,   T≈43.2 p
    """
    rate = max(float(rate), 1e-9)
    batch_size = max(1, int(batch_size))
    n_zones = max(1, int(n_zones))
    n_trades = max(1, int(n_trades))
    W = float(total_work) if total_work is not None else float(n_zones)
    W = max(W, 1e-9)
    # time to complete one zone (work units / (work per period))
    t_zone = W / (rate * n_zones)
    # handoff: continuous — next trade starts when upstream finishes this zone
    # (lag periods from discrete engine mapped as 0 extra wait beyond finish)
    _ = handoff_lag  # kept for API compat; train uses finish-to-start

    start: List[List[float]] = [[0.0] * n_zones for _ in range(n_trades)]
    finish: List[List[float]] = [[0.0] * n_zones for _ in range(n_trades)]

    for z in range(n_zones):
        for i in range(n_trades):
            cand = [0.0]
            if z > 0:
                cand.append(finish[i][z - 1])
            if i > 0:
                cand.append(finish[i - 1][z])
            # batch: only release every batch_size zones to downstream
            # for OPF batch=1, same as zone-by-zone above
            if batch_size > 1 and i > 0:
                # downstream may start zone z only after upstream finished
                # the batch containing z (last zone of batch)
                batch_end = min(n_zones - 1, ((z // batch_size) + 1) * batch_size - 1)
                # still use same-zone upstream for wagon clarity of OPF;
                # batch>1 would delay — keep simple OPF train for takt education
                pass
            ps = max(cand)
            pe = ps + t_zone
            start[i][z] = ps
            finish[i][z] = pe

    cells: List[TaktPlanCell] = []
    for i in range(n_trades):
        for z in range(n_zones):
            cells.append(
                TaktPlanCell(
                    trade_index=i,
                    zone=z + 1,
                    period_start=float(start[i][z]),
                    period_end=float(finish[i][z]),
                    planned_rate=rate * n_zones / W,  # zona/periode efektif
                )
            )

    duration = float(max(finish[i][n_zones - 1] for i in range(n_trades)))
    return TaktPlan(
        n_trades=n_trades,
        n_zones=n_zones,
        batch_size=batch_size,
        rate=rate * n_zones / W,
        cells=cells,
        duration=float(round(duration, 4)),
        takt_time=t_zone,
        handoff_lag=max(0, int(handoff_lag)),
    )




def apply_takt_buffers(
    n_trades: int,
    n_zones: int,
    base_rate: float,
    buffers: TaktBufferConfig,
) -> dict:
    """
    Compute effective rate, handoff lag, batch, and planned duration pads
    from base capacity + TPI-style buffers.
    """
    base_rate = max(float(base_rate), 1e-9)
    cap = max(0.0, float(buffers.independent_capacity))
    rate_eff = base_rate * (1.0 + cap)
    # Wagon buffer: extra process time per zone
    t_proc = 1.0 / rate_eff
    t_zone = t_proc + float(max(0, int(buffers.diagonal_wagon)))
    rate_plan = 1.0 / max(t_zone, 1e-9)
    # Sequence + buffer wagons → extra handoff lag (default engine lag = 1)
    lag = 1 + max(0, int(buffers.diagonal_sequence)) + max(0, int(buffers.diagonal_buffer_wagon))
    plan = build_takt_plan(
        n_trades=n_trades,
        n_zones=n_zones,
        batch_size=1,  # one-piece train; buffers absorb variation
        rate=rate_plan,
        handoff_lag=1,  # engine lag fixed; extra diagonal lag via duration pad
    )
    # Approximate diagonal lag effect: each extra lag unit adds ~n_trades-1 periods
    # (one between each consecutive pair) once per "wave" — use (n_trades-1)*extra
    extra_lag = max(0, lag - 1)
    diagonal_pad = extra_lag * max(0, n_trades - 1)
    vertical = max(0, int(buffers.vertical))
    horizontal = max(0, int(buffers.horizontal_end))
    duration_with_buffers = int(plan.duration) + diagonal_pad + vertical + horizontal
    return {
        "rate_plan": rate_plan,
        "rate_eff": rate_eff,
        "takt_time": t_zone,
        "batch": 1,
        "handoff_lag": lag,
        "plan": plan,
        "duration_work": int(plan.duration),
        "duration_plan": duration_with_buffers,
        "pad_diagonal": diagonal_pad,
        "pad_vertical": vertical,
        "pad_horizontal": horizontal,
        "buffers": buffers,
    }


def build_takt_plan_with_buffers(
    n_trades: int = 5,
    n_zones: int = 20,
    rate: float = 1.0,
    *,
    capacity_buffer: float = 0.0,
    time_buffer: int = 0,
    inventory_buffer: int = 0,
) -> "TaktPlan":
    """
    Takt plan with one-piece flow + three buffer types (lean construction).

    - **One-piece flow** always (batch_size = 1).
    - **Capacity buffer** (0..1+): extra capacity fraction.
      effective_rate = rate * (1 + capacity_buffer)
    - **Time buffer** (periods): extra lag after each handoff beyond the
      default next-period lag (engine lag is 1; we model extra by
      reducing effective rate so each zone takes process_time + time_buffer
      ... actually time buffer = added periods of lag between trades:
      simulated via batch=1 and staggered delay — here we add time_buffer
      to process time: t_e' = 1/rate + time_buffer, i.e. slower planned pace
      OR as pure lag: use inventory-style wait. We use **added process slack**:
      effective_rate = 1 / (1/rate + time_buffer) when time_buffer>0 and rate-based.
    - **Inventory buffer** (zones): minimum inbound zones before a trade
      starts / continues (decoupling stock). Modelled as batch_size =
      max(1, inventory_buffer) for release grouping while plan label stays OPF
      when inventory_buffer<=1.

    Primary design variable remains **n_zones**.
    """
    rate = max(float(rate), 1e-9)
    cap_b = max(0.0, float(capacity_buffer))
    time_b = max(0, int(time_buffer))
    inv_b = max(0, int(inventory_buffer))

    # Effective process rate after capacity buffer
    rate_eff = rate * (1.0 + cap_b)
    # Time buffer: add slack periods to each zone's process time
    # t = 1/rate_eff + time_b  → rate_plan = 1/t
    t_proc = 1.0 / rate_eff
    t_with_time_buf = t_proc + float(time_b)
    rate_plan = 1.0 / max(t_with_time_buf, 1e-9)

    # Inventory buffer: need inv_b zones released together (decoupling)
    # inv_b=0 or 1 → pure one-piece (batch 1); inv_b=2 → release every 2, etc.
    batch = 1 if inv_b <= 1 else inv_b

    plan = build_takt_plan(
        n_trades=n_trades,
        n_zones=n_zones,
        batch_size=batch,
        rate=rate_plan,
        handoff_lag=1,
    )
    # Annotate nominal rate (before buffers) on metadata via batch_size field kept
    # Store effective design in rate field as rate_plan; keep batch_size as used
    plan.rate = rate_plan
    # monkey-patch design attrs for UI (TaktPlan is dataclass — add if fields exist)
    return plan


def run_takt_simulation_with_buffers(
    n_trades: int,
    n_zones: int,
    rate: float,
    variability: str,
    seed: Optional[int],
    *,
    capacity_buffer: float = 0.0,
    time_buffer: int = 0,
    inventory_buffer: int = 0,
    pair_from_base_and_var=None,
) -> Tuple["ParadeResult", "TaktPlan"]:
    """
    Run zone-flow sim with OPF + buffers; return (result, ideal_plan).
    pair_from_base_and_var: optional callable(base, var)->pair from app.
    """
    from parade_of_trades_core import ParadeConfig, ParadeOfTrades

    rate = max(float(rate), 1e-9)
    cap_b = max(0.0, float(capacity_buffer))
    time_b = max(0, int(time_buffer))
    inv_b = max(0, int(inventory_buffer))

    rate_eff = rate * (1.0 + cap_b)
    t_proc = 1.0 / rate_eff
    rate_plan = 1.0 / max(t_proc + float(time_b), 1e-9)
    batch = 1 if inv_b <= 1 else inv_b

    plan = build_takt_plan(
        n_trades=n_trades,
        n_zones=n_zones,
        batch_size=batch,
        rate=rate_plan,
        handoff_lag=1,
    )

    if pair_from_base_and_var is not None:
        pairs = [pair_from_base_and_var(rate_plan, variability)] * n_trades
    else:
        # deterministic or simple no-var
        pairs = [(rate_plan, rate_plan, 0.5, rate_plan, True)] * n_trades

    # Build config manually if pairs are tuples of 5
    cfg = ParadeConfig.from_pairs(
        pairs,
        total_units=n_zones,
        seed=seed,
        zone_flow=True,
        batch_size=batch,
    )
    result = ParadeOfTrades(cfg).run()
    return result, plan


def required_rate_for_duration(
    n_trades: int,
    n_zones: int,
    batch_size: int,
    target_duration: int,
    handoff_lag: int = 1,
    rate_min: float = 0.2,
    rate_max: float = 5.0,
    tol: float = 0.01,
) -> dict:
    """
    Binary-search capacity rate (zona/periode) so ideal takt plan duration
    is <= target_duration. Returns feasibility and bounding rates.
    """
    target_duration = max(1, int(target_duration))
    lo, hi = float(rate_min), float(rate_max)
    best = None
    # check extremes
    p_lo = build_takt_plan(n_trades, n_zones, batch_size, lo, handoff_lag)
    p_hi = build_takt_plan(n_trades, n_zones, batch_size, hi, handoff_lag)
    if p_hi.duration > target_duration:
        return {
            "feasible": False,
            "rate": hi,
            "duration": p_hi.duration,
            "target": target_duration,
            "message": (
                f"Target {target_duration} terlalu ketat bahkan di rate {hi:g} "
                f"(rencana {p_hi.duration}). Naikkan target atau turunkan zona/batch."
            ),
        }
    # search
    for _ in range(40):
        mid = 0.5 * (lo + hi)
        p = build_takt_plan(n_trades, n_zones, batch_size, mid, handoff_lag)
        if p.duration <= target_duration:
            best = (mid, p.duration)
            hi = mid
        else:
            lo = mid
        if hi - lo < tol:
            break
    if best is None:
        best = (hi, p_hi.duration)
    rate, dur = best
    return {
        "feasible": True,
        "rate": rate,
        "duration": dur,
        "target": target_duration,
        "takt_time": 1.0 / rate,
        "message": (
            f"Butuh kapasitas ≈ {rate:.3f} zona/periode "
            f"(takt time ≈ {1.0/rate:.2f} periode/zona) untuk durasi ≤ {target_duration}."
        ),
    }


def takt_design_table(
    n_trades: int,
    n_zones: int,
    rates: Optional[Sequence[float]] = None,
    batches: Optional[Sequence[int]] = None,
) -> List[dict]:
    """What-if table: planned duration for combinations of rate × batch."""
    if rates is None:
        rates = [1.0 / 3.0, 0.5, 1.0, 2.0, 3.0]
    if batches is None:
        batches = [1, 2, 3, 4, 5]
    rows = []
    for r in rates:
        for b in batches:
            p = build_takt_plan(n_trades, n_zones, int(b), float(r), 1)
            rows.append({
                "Kapasitas": round(float(r), 4),
                "Batch / ukuran takt": int(b),
                "Takt time": round(p.takt_time, 3),
                "Durasi rencana": p.duration,
                "Zona": n_zones,
            })
    return rows


def takt_plan_reliability(
    result: ParadeResult,
    plan: TaktPlan,
) -> dict:
    """
    Compare simulated zone completions to takt plan.

    For each trade, approximate zone completion periods from cumulative
    history (first period cumulative >= z). Reliability = share of
    (trade, zone) finished on or before plan.period_end.
    """
    n = result.config.n_trades
    n_zones = min(plan.n_zones, result.config.total_units)
    # actual finish period for trade i zone z (1-based z)
    actual: List[List[Optional[int]]] = [
        [None] * n_zones for _ in range(n)
    ]
    for rec in result.history:
        for i in range(n):
            cum = int(rec.cumulative[i])
            for z in range(min(cum, n_zones)):
                if actual[i][z] is None:
                    actual[i][z] = rec.period

    on_time = 0
    total = 0
    late = 0
    early = 0
    detail = []
    plan_by = {(c.trade_index, c.zone): c for c in plan.cells}
    for i in range(n):
        for z in range(n_zones):
            cell = plan_by.get((i, z + 1))
            if cell is None:
                continue
            act = actual[i][z]
            total += 1
            if act is None:
                late += 1
                status = "belum"
            elif act <= cell.period_end:
                on_time += 1
                status = "tepat/awal"
                if act < cell.period_start:
                    early += 1
            else:
                late += 1
                status = "telat"
            detail.append({
                "Tim": i + 1,
                "Zona": z + 1,
                "Rencana selesai": cell.period_end,
                "Aktual selesai": act if act is not None else "—",
                "Status": status,
            })
    rel = on_time / total if total else 0.0
    return {
        "reliability": rel,
        "on_time": on_time,
        "late": late,
        "early": early,
        "total": total,
        "plan_duration": plan.duration,
        "actual_duration": result.duration,
        "detail": detail,
    }


# ---------------------------------------------------------------------------
# Statistics helpers
# ---------------------------------------------------------------------------

def _percentile(sorted_vals: Sequence[float], p: float) -> float:
    """Linear-interpolation percentile; ``p`` in [0, 100]."""
    if not sorted_vals:
        return float("nan")
    if len(sorted_vals) == 1:
        return float(sorted_vals[0])
    k = (len(sorted_vals) - 1) * (p / 100.0)
    f = math.floor(k)
    c = math.ceil(k)
    if f == c:
        return float(sorted_vals[int(k)])
    d0 = sorted_vals[f] * (c - k)
    d1 = sorted_vals[c] * (k - f)
    return float(d0 + d1)


@dataclass
class MetricStats:
    """Descriptive stats for one scalar metric across replications."""

    name: str
    n: int
    mean: float
    std: float
    min: float
    p25: float
    median: float
    p75: float
    max: float

    @classmethod
    def from_values(cls, name: str, values: Sequence[float]) -> "MetricStats":
        vals = [float(v) for v in values]
        if not vals:
            return cls(name, 0, float("nan"), float("nan"), float("nan"),
                       float("nan"), float("nan"), float("nan"), float("nan"))
        s = sorted(vals)
        std = statistics.stdev(vals) if len(vals) > 1 else 0.0
        return cls(
            name=name,
            n=len(vals),
            mean=statistics.mean(vals),
            std=std,
            min=s[0],
            p25=_percentile(s, 25),
            median=_percentile(s, 50),
            p75=_percentile(s, 75),
            max=s[-1],
        )

    def as_dict(self) -> dict:
        return {
            "metric": self.name,
            "n": self.n,
            "mean": self.mean,
            "std": self.std,
            "min": self.min,
            "p25": self.p25,
            "median": self.median,
            "p75": self.p75,
            "max": self.max,
        }


# ---------------------------------------------------------------------------
# Replication batch
# ---------------------------------------------------------------------------

@dataclass
class ReplicationBatch:
    """Results of many independent runs of the same configuration."""

    config: ParadeConfig
    n_reps: int
    seed_base: int
    results: List[ParadeResult] = field(default_factory=list)
    seeds: List[int] = field(default_factory=list)

    # -- derived series -----------------------------------------------------

    @property
    def durations(self) -> List[int]:
        return [r.duration for r in self.results]

    @property
    def throughputs(self) -> List[float]:
        return [r.system_throughput for r in self.results]

    @property
    def total_idles(self) -> List[int]:
        return [r.total_idle_capacity for r in self.results]

    @property
    def peak_wips(self) -> List[int]:
        out = []
        for r in self.results:
            peak = max((sum(h.buffers) for h in r.history), default=0)
            out.append(peak)
        return out

    @property
    def standby_totals(self) -> List[int]:
        return [r.total_standby_used for r in self.results]

    def time_on_site_by_trade(self) -> List[List[int]]:
        """series[trade_idx] = list of time_on_site across reps."""
        n = self.config.n_trades
        series: List[List[int]] = [[] for _ in range(n)]
        for r in self.results:
            for i, m in enumerate(r.trade_metrics):
                series[i].append(m.time_on_site)
        return series

    def idle_by_trade(self) -> List[List[int]]:
        n = self.config.n_trades
        series: List[List[int]] = [[] for _ in range(n)]
        for r in self.results:
            for i, m in enumerate(r.trade_metrics):
                series[i].append(m.total_idle)
        return series

    # -- stats --------------------------------------------------------------

    def stats(self) -> Dict[str, MetricStats]:
        """System-level metric statistics."""
        return {
            "duration": MetricStats.from_values("duration", self.durations),
            "throughput": MetricStats.from_values("throughput", self.throughputs),
            "total_idle": MetricStats.from_values("total_idle", self.total_idles),
            "peak_wip": MetricStats.from_values("peak_wip", self.peak_wips),
            "total_standby": MetricStats.from_values(
                "total_standby", self.standby_totals
            ),
        }

    def trade_time_on_site_stats(self) -> List[MetricStats]:
        series = self.time_on_site_by_trade()
        out = []
        for i, vals in enumerate(series):
            name = self.config.trades[i].name
            out.append(MetricStats.from_values(f"time_on_site[{name}]", vals))
        return out

    def summary_table(self) -> List[dict]:
        """Rows suitable for display / DataFrame."""
        return [s.as_dict() for s in self.stats().values()]

    def print_summary(self) -> None:
        sep = "=" * 72
        print(sep)
        print("PARADE OF TRADES – Replication Summary")
        print(sep)
        print(f"  n_reps    : {self.n_reps}")
        print(f"  seed_base : {self.seed_base}")
        print(f"  mode      : {self.config.mode_label()}")
        pairs = ", ".join(t.label() for t in self.config.trades)
        print(f"  capacities: [{pairs}]")
        print(f"  units     : {self.config.total_units}")
        print("-" * 72)
        hdr = (
            f"{'Metric':<16}  {'Mean':>8}  {'Std':>8}  {'Min':>7}  "
            f"{'P25':>7}  {'Med':>7}  {'P75':>7}  {'Max':>7}"
        )
        print(hdr)
        print("-" * 72)
        for s in self.stats().values():
            print(
                f"{s.name:<16}  {s.mean:>8.2f}  {s.std:>8.2f}  {s.min:>7.1f}  "
                f"{s.p25:>7.1f}  {s.median:>7.1f}  {s.p75:>7.1f}  {s.max:>7.1f}"
            )
        print("-" * 72)
        print("  Time on site by trade:")
        for s in self.trade_time_on_site_stats():
            short = s.name.replace("time_on_site[", "").rstrip("]")
            if len(short) > 22:
                short = short[:21] + "…"
            print(
                f"    {short:<22}  mean={s.mean:6.2f}  std={s.std:5.2f}  "
                f"[{s.min:.0f} .. {s.max:.0f}]"
            )
        print(sep)

    # -- export -------------------------------------------------------------

    def export_csv(self, path: PathLike) -> Path:
        """One row per replication (system metrics + per-trade time on site)."""
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        fieldnames = [
            "rep", "seed", "duration", "throughput", "total_idle",
            "peak_wip", "total_standby",
        ]
        for i in range(self.config.n_trades):
            fieldnames.append(f"time_on_site_t{i + 1}")
            fieldnames.append(f"idle_t{i + 1}")
            fieldnames.append(f"util_t{i + 1}")

        with path.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            for k, (seed, r) in enumerate(zip(self.seeds, self.results)):
                peak = max((sum(h.buffers) for h in r.history), default=0)
                row = {
                    "rep": k + 1,
                    "seed": seed,
                    "duration": r.duration,
                    "throughput": r.system_throughput,
                    "total_idle": r.total_idle_capacity,
                    "peak_wip": peak,
                    "total_standby": r.total_standby_used,
                }
                for i, m in enumerate(r.trade_metrics):
                    row[f"time_on_site_t{i + 1}"] = m.time_on_site
                    row[f"idle_t{i + 1}"] = m.total_idle
                    row[f"util_t{i + 1}"] = round(m.utilization, 4)
                w.writerow(row)
        return path

    def export_excel(self, path: PathLike) -> Path:
        """Excel workbook: Summary | Replications | Config."""
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError as e:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl"
            ) from e

        wb = openpyxl.Workbook()

        # --- Summary ---
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Parade of Trades – Replication Summary"])
        ws["A1"].font = Font(bold=True, size=12)
        ws.append([])
        ws.append(["n_reps", self.n_reps])
        ws.append(["seed_base", self.seed_base])
        ws.append(["mode", self.config.mode_label()])
        ws.append(["total_units", self.config.total_units])
        ws.append(["takt_rate", self.config.takt_rate])
        ws.append(["standby_capacity", self.config.standby_capacity])
        ws.append(["staggered", self.config.staggered_mobilization])
        ws.append([])
        ws.append(
            ["metric", "n", "mean", "std", "min", "p25", "median", "p75", "max"]
        )
        for s in self.stats().values():
            ws.append(
                [s.name, s.n, s.mean, s.std, s.min, s.p25, s.median, s.p75, s.max]
            )
        ws.append([])
        ws.append(["Time on site by trade"])
        ws.append(
            ["trade", "n", "mean", "std", "min", "p25", "median", "p75", "max"]
        )
        for s in self.trade_time_on_site_stats():
            ws.append(
                [s.name, s.n, s.mean, s.std, s.min, s.p25, s.median, s.p75, s.max]
            )

        # --- Replications ---
        ws2 = wb.create_sheet("Replications")
        headers = [
            "rep", "seed", "duration", "throughput", "total_idle",
            "peak_wip", "total_standby",
        ]
        for i, t in enumerate(self.config.trades):
            headers += [
                f"tos_t{i + 1}",
                f"idle_t{i + 1}",
                f"util_t{i + 1}",
                f"stby_t{i + 1}",
            ]
        ws2.append(headers)
        for k, (seed, r) in enumerate(zip(self.seeds, self.results)):
            peak = max((sum(h.buffers) for h in r.history), default=0)
            row = [
                k + 1, seed, r.duration, r.system_throughput,
                r.total_idle_capacity, peak, r.total_standby_used,
            ]
            for m in r.trade_metrics:
                row += [
                    m.time_on_site, m.total_idle,
                    round(m.utilization, 4), m.total_standby_used,
                ]
            ws2.append(row)

        # --- Config ---
        ws3 = wb.create_sheet("Config")
        ws3.append(["#", "name", "low", "high", "mean"])
        for i, t in enumerate(self.config.trades):
            ws3.append([i + 1, t.name, t.low, t.high, t.mean])

        wb.save(path)
        return path


def run_replications(
    config: ParadeConfig,
    n_reps: int = 100,
    seed_base: int = 0,
    verbose: bool = False,
) -> ReplicationBatch:
    """
    Run ``n_reps`` independent simulations.

    Seeds used: seed_base, seed_base+1, …, seed_base+n_reps-1.
    Each replication gets a fresh ``ParadeConfig`` clone with its own seed.
    """
    if n_reps <= 0:
        raise ValueError("n_reps must be positive")

    batch = ReplicationBatch(
        config=config, n_reps=n_reps, seed_base=seed_base
    )
    for k in range(n_reps):
        seed = seed_base + k
        cfg = ParadeConfig(
            trades=list(config.trades),
            total_units=config.total_units,
            seed=seed,
            takt_rate=config.takt_rate,
            standby_capacity=config.standby_capacity,
            same_period_handoff=config.same_period_handoff,
            staggered_mobilization=config.staggered_mobilization,
        )
        sim = ParadeOfTrades(cfg)
        result = sim.run()
        batch.results.append(result)
        batch.seeds.append(seed)

    if verbose:
        batch.print_summary()
    return batch


# ---------------------------------------------------------------------------
# Multi-scenario comparison across replications
# ---------------------------------------------------------------------------

@dataclass
class ScenarioComparison:
    """Named replication batches for side-by-side comparison."""

    batches: Dict[str, ReplicationBatch]

    def summary_rows(self) -> List[dict]:
        rows = []
        for name, batch in self.batches.items():
            st = batch.stats()
            rows.append(
                {
                    "scenario": name,
                    "n_reps": batch.n_reps,
                    "mode": batch.config.mode_label(),
                    "duration_mean": st["duration"].mean,
                    "duration_std": st["duration"].std,
                    "duration_min": st["duration"].min,
                    "duration_max": st["duration"].max,
                    "throughput_mean": st["throughput"].mean,
                    "idle_mean": st["total_idle"].mean,
                    "peak_wip_mean": st["peak_wip"].mean,
                    "standby_mean": st["total_standby"].mean,
                }
            )
        return rows

    def print_summary(self) -> None:
        sep = "=" * 88
        print(sep)
        print("PARADE OF TRADES – Multi-Scenario Replication Comparison")
        print(sep)
        hdr = (
            f"{'Scenario':<22}  {'N':>4}  {'Dur μ':>7}  {'Dur σ':>7}  "
            f"{'Dur min':>7}  {'Dur max':>7}  {'Idle μ':>8}  {'WIP μ':>7}"
        )
        print(hdr)
        print("-" * 88)
        for row in self.summary_rows():
            print(
                f"{row['scenario']:<22}  {row['n_reps']:>4}  "
                f"{row['duration_mean']:>7.2f}  {row['duration_std']:>7.2f}  "
                f"{row['duration_min']:>7.0f}  {row['duration_max']:>7.0f}  "
                f"{row['idle_mean']:>8.1f}  {row['peak_wip_mean']:>7.1f}"
            )
        print(sep)

    def export_excel(self, path: PathLike) -> Path:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError as e:
            raise ImportError("openpyxl required for Excel export") from e

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Comparison"
        ws.append(["Parade of Trades – Scenario Comparison"])
        ws["A1"].font = Font(bold=True, size=12)
        ws.append([])
        headers = list(self.summary_rows()[0].keys()) if self.batches else []
        if headers:
            ws.append(headers)
            for row in self.summary_rows():
                ws.append([row[h] for h in headers])

        for name, batch in self.batches.items():
            safe = name[:28].replace("/", "-")
            ws_b = wb.create_sheet(safe)
            st = batch.stats()
            ws_b.append(
                ["metric", "n", "mean", "std", "min", "p25", "median", "p75", "max"]
            )
            for s in st.values():
                ws_b.append(
                    [s.name, s.n, s.mean, s.std, s.min, s.p25, s.median, s.p75, s.max]
                )
            ws_b.append([])
            ws_b.append(["Time on site by trade"])
            for s in batch.trade_time_on_site_stats():
                ws_b.append(
                    [s.name, s.n, s.mean, s.std, s.min, s.p25, s.median, s.p75, s.max]
                )

        wb.save(path)
        return path


def compare_scenarios(
    configs: Dict[str, ParadeConfig],
    n_reps: int = 100,
    seed_base: int = 0,
    verbose: bool = True,
) -> ScenarioComparison:
    """Run replications for each named config and return a comparison object."""
    batches: Dict[str, ReplicationBatch] = {}
    for name, cfg in configs.items():
        batches[name] = run_replications(
            cfg, n_reps=n_reps, seed_base=seed_base, verbose=False
        )
    cmp = ScenarioComparison(batches=batches)
    if verbose:
        cmp.print_summary()
    return cmp



def takt_plan_from_lob_result(result: ParadeResult) -> "TaktPlan":
    """
    Translate a classic Parade LOB run into a TaktPlan-like schedule.

    For each trade and each completed unit (1..N), estimate the period when
    that unit was finished from the cumulative series (first period cum >= z).
    Start period ≈ previous unit finish + 1 (or first production period).
    """
    n = result.config.n_trades
    n_units = int(result.config.total_units)
    cum = result.cumulative_series()  # series[i][p] with p=0 → 0

    # actual finish period for trade i, unit z (0-based unit index)
    finish: List[List[Optional[int]]] = [[None] * n_units for _ in range(n)]
    start: List[List[Optional[int]]] = [[None] * n_units for _ in range(n)]

    for i in range(n):
        series = cum[i]
        # series index = period
        for p in range(1, len(series)):
            prev = int(series[p - 1])
            cur = int(series[p])
            if cur > prev:
                for z in range(prev, min(cur, n_units)):
                    if finish[i][z] is None:
                        finish[i][z] = p
                        if z == 0:
                            start[i][z] = p  # first unit of this trade that period
                        else:
                            prev_f = finish[i][z - 1]
                            start[i][z] = (prev_f + 1) if prev_f is not None and prev_f < p else p

    cells: List[TaktPlanCell] = []
    for i in range(n):
        for z in range(n_units):
            pe = finish[i][z] if finish[i][z] is not None else result.duration
            ps = start[i][z] if start[i][z] is not None else pe
            if ps > pe:
                ps = pe
            cells.append(
                TaktPlanCell(
                    trade_index=i,
                    zone=z + 1,
                    period_start=int(ps),
                    period_end=int(pe),
                    planned_rate=float(result.config.takt_rate or (result.config.trades[i].mean or 1.0)),
                )
            )

    rate = float(result.config.takt_rate) if result.config.takt_enabled else float(
        result.config.trades[0].mean or 1.0
    )
    return TaktPlan(
        n_trades=n,
        n_zones=n_units,
        batch_size=1,
        rate=rate,
        cells=cells,
        duration=int(result.duration),
        takt_time=(1.0 / rate) if rate > 0 else 1.0,
        handoff_lag=0 if result.config.same_period_handoff else 1,
    )


def tommelein_run_metrics(result: ParadeResult) -> dict:
    """Summary row for one Tommelein (2020) scenario run."""
    cfg = result.config
    t0 = cfg.trades[0]
    mean_die = float(t0.mean) if t0.mean else (float(t0.low) + float(t0.high)) / 2.0
    # periods meeting takt (when takt enabled): production >= takt when work available
    periods_ok = 0
    periods_work = 0
    takt = cfg.takt_rate
    if takt is not None:
        for rec in result.history:
            for i in range(cfg.n_trades):
                # count only when trade had capacity roll (mobilized / working)
                if rec.capacity[i] and rec.capacity[i] > 0:
                    periods_work += 1
                    if rec.production[i] + 1e-9 >= min(takt, rec.effective_capacity[i]):
                        # met commitment if produced at least takt when possible
                        if rec.effective_capacity[i] >= takt - 1e-9 and rec.production[i] + 1e-9 >= takt:
                            periods_ok += 1
                        elif rec.effective_capacity[i] < takt and rec.production[i] + 1e-9 >= rec.effective_capacity[i] - 1e-9:
                            periods_ok += 1  # limited by work not capacity
    rel = (periods_ok / periods_work) if periods_work else None
    return {
        "Dadu": f"{int(t0.low)}–{int(t0.high)}" if t0.low != t0.high else str(int(t0.low)),
        "Mean dadu": f"{mean_die:.2f}",
        "Takt": str(cfg.takt_rate) if cfg.takt_rate is not None else "—",
        "Standby": str(cfg.standby_capacity) if cfg.takt_enabled else "—",
        "Durasi": int(result.duration),
        "TH": round(float(result.system_throughput), 3),
        "Idle": float(result.total_idle_capacity),
        "Standby dipakai": int(result.total_standby_used),
        "Reliability takt": f"{100 * rel:.0f}%" if rel is not None else "—",
    }


def compare_tommelein2020(

    n_reps: int = 100,
    seed_base: int = 0,
    total_units: int = 100,
    staggered: bool = False,
    same_period_handoff: bool = False,
    verbose: bool = True,
) -> ScenarioComparison:
    """Replicate the three scenarios from Tommelein (2020)."""
    configs = tommelein2020_scenarios(
        total_units=total_units,
        seed=None,
        staggered=staggered,
        same_period_handoff=same_period_handoff,
    )
    return compare_scenarios(
        configs, n_reps=n_reps, seed_base=seed_base, verbose=verbose
    )


# ---------------------------------------------------------------------------
# Single-result export
# ---------------------------------------------------------------------------

def export_result_csv(
    result: ParadeResult,
    path: PathLike,
    *,
    include_history: bool = True,
) -> Path:
    """
    Export a single run.

    If ``include_history``, writes period-level history CSV.
    Always also writes a sibling ``*_summary.csv`` with trade metrics.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    # Summary trades
    summary_path = path.with_name(path.stem + "_summary.csv")
    with summary_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["field", "value"])
        w.writerow(["seed", result.config.seed])
        w.writerow(["total_units", result.config.total_units])
        w.writerow(["mode", result.config.mode_label()])
        w.writerow(["duration", result.duration])
        w.writerow(["throughput", result.system_throughput])
        w.writerow(["ideal_duration", result.ideal_duration])
        w.writerow(["total_idle", result.total_idle_capacity])
        w.writerow(["total_standby", result.total_standby_used])
        w.writerow([])
        w.writerow(
            [
                "trade", "low", "high", "executions", "production", "idle",
                "utilization", "finish", "time_on_site", "standby_used",
            ]
        )
        for m in result.trade_metrics:
            w.writerow(
                [
                    m.name, m.capacity_pair[0], m.capacity_pair[1],
                    m.executions, m.total_production, m.total_idle,
                    round(m.utilization, 4), m.periods_to_finish,
                    m.time_on_site, m.total_standby_used,
                ]
            )
        if result.max_buffer:
            w.writerow([])
            w.writerow(["buffer", "max_wip"])
            for j, mx in enumerate(result.max_buffer):
                w.writerow([j + 1, mx])

    if include_history:
        sim_rows = _history_dicts(result)
        if sim_rows:
            with path.open("w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=list(sim_rows[0].keys()))
                w.writeheader()
                w.writerows(sim_rows)
        else:
            path.write_text("", encoding="utf-8")

    return path


def export_result_excel(result: ParadeResult, path: PathLike) -> Path:
    """Excel workbook for a single run: Summary | Trades | History | Buffers."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError as e:
        raise ImportError("openpyxl required for Excel export") from e

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Parade of Trades – Run Summary"])
    ws["A1"].font = Font(bold=True, size=12)
    peak = max((sum(h.buffers) for h in result.history), default=0)
    rows = [
        ("seed", result.config.seed),
        ("total_units", result.config.total_units),
        ("mode", result.config.mode_label()),
        ("takt_rate", result.config.takt_rate),
        ("standby_capacity", result.config.standby_capacity),
        ("staggered_mobilization", result.config.staggered_mobilization),
        ("duration", result.duration),
        ("ideal_duration", result.ideal_duration),
        ("throughput", result.system_throughput),
        ("total_idle", result.total_idle_capacity),
        ("total_standby", result.total_standby_used),
        ("peak_simultaneous_wip", peak),
    ]
    ws.append([])
    for k, v in rows:
        ws.append([k, v])

    ws2 = wb.create_sheet("Trades")
    ws2.append(
        [
            "#", "name", "low", "high", "executions", "production", "idle",
            "utilization", "finish", "time_on_site", "standby_used",
            "effective_capacity",
        ]
    )
    for i, m in enumerate(result.trade_metrics):
        ws2.append(
            [
                i + 1, m.name, m.capacity_pair[0], m.capacity_pair[1],
                m.executions, m.total_production, m.total_idle,
                m.utilization, m.periods_to_finish, m.time_on_site,
                m.total_standby_used, m.total_effective_capacity,
            ]
        )

    ws3 = wb.create_sheet("History")
    hist = _history_dicts(result)
    if hist:
        ws3.append(list(hist[0].keys()))
        for row in hist:
            ws3.append(list(row.values()))

    ws4 = wb.create_sheet("Buffers")
    ws4.append(["interface", "from", "to", "max_wip"])
    for j, mx in enumerate(result.max_buffer):
        up = result.config.trades[j].name
        down = result.config.trades[j + 1].name
        ws4.append([j + 1, up, down, mx])

    # Analysis sheets for presentation
    try:
        ll = littles_law_metrics(result)
        d = littles_operations_curve(result)
        ws5 = wb.create_sheet("Littles_Law")
        ws5.append(["metric", "value"])
        for k, v in [
            ("TH", ll.throughput),
            ("WIP_pipeline_avg", ll.avg_pipeline_wip),
            ("CT_pipeline", ll.cycle_time_pipeline),
            ("WIP_buffer_avg", ll.avg_buffer_wip),
            ("CT_buffer", ll.cycle_time_buffer),
            ("TH_x_CT_check", ll.check_pipeline),
            ("W_min", d.get("w_min")),
            ("W_opt", d.get("w_opt")),
            ("CONWIP_suggest", d.get("conwip")),
            ("TH_max", d.get("th_max")),
            ("T0", d.get("t0")),
            ("V_factor", d.get("v_factor", d.get("v"))),
        ]:
            ws5.append([k, v])
    except Exception:
        pass

    try:
        kg = kingman_metrics(result)
        comb = kingman_combined(result)
        ws6 = wb.create_sheet("Kingman")
        ws6.append(["metric", "value"])
        ws6.append(["u_bar", comb.get("u_bar")])
        ws6.append(["V", comb.get("v")])
        ws6.append(["sum_ct_kingman", getattr(kg, "sum_ct_kingman", None)])
        ws6.append(["sum_ct_observed", getattr(kg, "sum_ct_observed", None)])
        ws6.append(["system_ct_little", getattr(kg, "system_ct_little", None)])
        ws6.append([])
        rows = kg.as_rows() if hasattr(kg, "as_rows") else []
        if rows:
            ws6.append(list(rows[0].keys()))
            for row in rows:
                ws6.append(list(row.values()))
    except Exception:
        pass

    try:
        fr = inventory_fill_rate_metrics(result)
        ws7 = wb.create_sheet("Inventory_FR")
        ws7.append(["metric", "value"])
        for k, v in fr.items():
            if not isinstance(v, (list, dict)):
                ws7.append([k, v])
        interfaces = fr.get("interfaces") or []
        if interfaces:
            ws7.append([])
            ws7.append(list(interfaces[0].keys()))
            for row in interfaces:
                ws7.append(list(row.values()))
    except Exception:
        pass

    # LOB cumulative table
    try:
        ws8 = wb.create_sheet("LOB_cumulative")
        n = result.config.n_trades
        header = ["period"] + [f"T{i+1}_cum" for i in range(n)] + [f"B{j+1}" for j in range(max(0, n - 1))]
        ws8.append(header)
        for rec in result.history:
            row = [rec.period] + list(rec.cumulative)
            row += list(rec.buffers)
            ws8.append(row)
    except Exception:
        pass

    wb.save(path)
    return path


def export_comparison_excel(
    results: Dict[str, ParadeResult],
    path: PathLike,
    meta: Optional[Dict[str, dict]] = None,
) -> Path:
    """Multi-scenario workbook for presentation."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ringkasan"
    ws.append(["Parade Tim Kerja — Perbandingan Skenario"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])
    ws.append([
        "Skenario", "Durasi", "Ideal", "TH", "Idle", "Puncak_WIP",
        "WIP_avg", "CT", "W_min", "W_opt", "Batch", "Zona", "Seed",
        "Variability", "Pace",
    ])
    for name, r in results.items():
        ll = littles_law_metrics(r)
        try:
            d = littles_operations_curve(r)
            wmin, wopt = d["w_min"], d["w_opt"]
        except Exception:
            wmin = wopt = ""
        peak = max((sum(h.buffers) for h in r.history), default=0)
        m = (meta or {}).get(name, {})
        ws.append([
            name, r.duration, r.ideal_duration, round(ll.throughput, 4),
            r.total_idle_capacity, peak, round(ll.avg_pipeline_wip, 3),
            round(ll.cycle_time_pipeline, 3), wmin, wopt,
            r.config.batch_size, r.config.total_units, r.config.seed,
            m.get("var_label", ""), m.get("pace", ""),
        ])

    wu = wb.create_sheet("Utilisasi")
    wu.append(["Skenario", "Tim", "Produksi", "Idle", "Utilisasi"])
    for name, r in results.items():
        for tm in r.trade_metrics:
            wu.append([name, tm.name, tm.total_production, tm.total_idle, round(tm.utilization, 4)])

    wlob = wb.create_sheet("LOB_tim_terakhir")
    max_len = max((len(r.history) for r in results.values()), default=0)
    wlob.append(["period"] + list(results.keys()))
    for p in range(max_len + 1):
        row = [p]
        for r in results.values():
            cum = r.cumulative_series()[-1]
            row.append(cum[p] if p < len(cum) else "")
        wlob.append(row)

    for idx, (name, r) in enumerate(results.items(), 1):
        safe = re.sub(r"[^\w\-]+", "_", name)[:20]
        title = f"S{idx}_{safe}"[:31]
        wh = wb.create_sheet(title)
        hist = _history_dicts(r)
        if hist:
            wh.append(list(hist[0].keys()))
            for row in hist:
                wh.append(list(row.values()))

    wb.save(path)
    return path


def export_comparison_csv(
    results: Dict[str, ParadeResult],
    path: PathLike,
    meta: Optional[Dict[str, dict]] = None,
) -> Path:
    """Summary CSV of multi-scenario comparison."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "Skenario", "Durasi", "Ideal", "TH", "Idle", "Puncak_WIP",
            "WIP_avg", "CT", "W_min", "W_opt", "Batch", "Zona", "Seed",
            "Variability", "Pace",
        ])
        for name, r in results.items():
            ll = littles_law_metrics(r)
            try:
                d = littles_operations_curve(r)
                wmin, wopt = d["w_min"], d["w_opt"]
            except Exception:
                wmin = wopt = ""
            peak = max((sum(h.buffers) for h in r.history), default=0)
            m = (meta or {}).get(name, {})
            w.writerow([
                name, r.duration, r.ideal_duration, round(ll.throughput, 4),
                r.total_idle_capacity, peak, round(ll.avg_pipeline_wip, 3),
                round(ll.cycle_time_pipeline, 3), wmin, wopt,
                r.config.batch_size, r.config.total_units, r.config.seed,
                m.get("var_label", ""), m.get("pace", ""),
            ])
    return path


def export_takt_excel(
    result: ParadeResult,
    plan: "TaktPlan",
    path: PathLike,
    *,
    reliability: Optional[dict] = None,
    duration_plan: Optional[int] = None,
    buffers_info: Optional[dict] = None,
) -> Path:
    """Takt plan + actual run workbook."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    # base single-run workbook
    export_result_excel(result, path)
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.load_workbook(path)

    ws = wb.create_sheet("Takt_rencana", 0)
    ws.append(["Parade Tim Kerja — Takt Plan"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])
    ws.append(["zona", plan.n_zones])
    ws.append(["rate", plan.rate])
    ws.append(["takt_time", plan.takt_time])
    ws.append(["batch", plan.batch_size])
    ws.append(["durasi_kerja_rencana", plan.duration])
    if duration_plan is not None:
        ws.append(["durasi_rencana_plus_buffer", duration_plan])
    if buffers_info:
        for k, v in buffers_info.items():
            ws.append([f"buffer_{k}", v])
    if reliability:
        ws.append([])
        ws.append(["reliability", reliability.get("reliability")])
        ws.append(["durasi_aktual", reliability.get("actual_duration")])
        ws.append(["durasi_plan_cells", reliability.get("plan_duration")])

    ws2 = wb.create_sheet("Takt_cells")
    ws2.append(["trade", "zone", "period_start", "period_end", "planned_rate"])
    for c in plan.cells:
        ws2.append([c.trade_index + 1, c.zone, c.period_start, c.period_end, c.planned_rate])

    if reliability and reliability.get("detail"):
        ws3 = wb.create_sheet("Reliability_detail")
        det = reliability["detail"]
        if det:
            ws3.append(list(det[0].keys()))
            for row in det:
                ws3.append(list(row.values()))

    wb.save(path)
    return path


def export_takt_csv(
    result: ParadeResult,
    plan: "TaktPlan",
    path: PathLike,
    *,
    reliability: Optional[dict] = None,
) -> Path:
    """CSV: takt cells + optional reliability summary header via sibling files."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["trade", "zone", "period_start", "period_end", "planned_rate"])
        for c in plan.cells:
            w.writerow([c.trade_index + 1, c.zone, c.period_start, c.period_end, c.planned_rate])
    # also dump actual history
    export_result_csv(result, path.with_name(path.stem + "_history.csv"), include_history=True)
    if reliability:
        with path.with_name(path.stem + "_summary.csv").open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["metric", "value"])
            w.writerow(["reliability", reliability.get("reliability")])
            w.writerow(["actual_duration", reliability.get("actual_duration")])
            w.writerow(["plan_duration", reliability.get("plan_duration")])
            w.writerow(["plan_rate", plan.rate])
            w.writerow(["plan_zones", plan.n_zones])
    return path


def _history_dicts(result: ParadeResult) -> List[dict]:
    """Period-level rows from a ParadeResult (no live sim needed)."""
    rows: List[dict] = []
    n = result.config.n_trades
    for rec in result.history:
        row: dict = {"period": rec.period, "raw_remaining": rec.raw_remaining}
        for i in range(n):
            row[f"cap_{i + 1}"] = rec.capacity[i]
            row[f"prod_{i + 1}"] = rec.production[i]
            row[f"idle_{i + 1}"] = rec.idle_capacity[i]
            row[f"cum_{i + 1}"] = rec.cumulative[i]
            if rec.effective_capacity:
                row[f"eff_{i + 1}"] = rec.effective_capacity[i]
            if rec.standby_used:
                row[f"stby_{i + 1}"] = rec.standby_used[i]
        for j, b in enumerate(rec.buffers):
            row[f"buffer_{j + 1}"] = b
        rows.append(row)
    return rows


# ---------------------------------------------------------------------------
# CLI demo
# ---------------------------------------------------------------------------

def _demo() -> None:
    print("\n>>> Tommelein 2020 scenarios — 50 replications each (seed_base=0)\n")
    cmp = compare_tommelein2020(n_reps=50, seed_base=0, verbose=True)

    out = Path("output")
    out.mkdir(exist_ok=True)
    xlsx = cmp.export_excel(out / "tommelein2020_comparison.xlsx")
    print(f"  wrote {xlsx}")

    # Single takt run export
    from parade_of_trades_core import run_preset

    r = run_preset(
        "low", seed=42, verbose=True,
        takt_rate=5, standby_capacity=1,
    )
    export_result_excel(r, out / "takt_run_example.xlsx")
    export_result_csv(r, out / "takt_run_history.csv")
    print("  wrote output/takt_run_example.xlsx and takt_run_history.csv")

    # Classic replications
    cfg = ParadeConfig.from_preset("medium", seed=None)
    batch = run_replications(cfg, n_reps=30, seed_base=100, verbose=True)
    batch.export_excel(out / "medium_reps.xlsx")
    batch.export_csv(out / "medium_reps.csv")
    print("  wrote output/medium_reps.xlsx / .csv")


if __name__ == "__main__":
    _demo()
